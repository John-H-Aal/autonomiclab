"""Excel export for Valsalva and Deep Breathing analysis results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from autonomiclab.analysis.deep_breathing import DeepBreathingResult
from autonomiclab.analysis.valsalva import ValsalvaResult
from autonomiclab.utils.logger import get_logger

log = get_logger(__name__)

_THIN = Side(style="thin", color="AAAAAA")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class ExcelExporter:
    """Write analysis results to formatted Excel workbooks."""

    # ── Valsalva ─────────────────────────────────────────────────────────────

    def export_valsalva(
        self, result: ValsalvaResult, output_dir: Path, mode: str = "auto"
    ) -> Path:
        """Write Valsalva results to a timestamped, mode-tagged xlsx file."""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = output_dir / f"{output_dir.parent.name}_valsalva_results_{stamp}_{mode}.xlsx"

        hdr_fill  = PatternFill("solid", start_color="1F4E79")
        sect_fill = PatternFill("solid", start_color="D6E4F0")
        warn_fill = PatternFill("solid", start_color="FFF2CC")
        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        sect_font = Font(name="Arial", bold=True, color="1F4E79", size=10)
        val_font  = Font(name="Arial", size=10)

        data: dict[str, Optional[float]] = {
            "t_baseline_start":         result.t_bl_s,
            "t_baseline_end":           result.t_bl_e,
            "t_S1_start":               result.t_S1s,
            "t_S1_end":                 result.t_S1e,
            "t_S2early_end":            result.t_S2es,
            "t_S2late_end_PIII_start":  result.t_S3s,
            "t_S3_end_PIV_start":       result.t_S3e,
            "t_S4_end":                 result.t_S4e,
            "SBP_baseline_mmHg":        result.avg_sbp,
            "SBP_IIe_nadir_mmHg":       result.v_nadir,
            "SBP_S2late_max_mmHg":      result.v_S2lmax,
            "SBP_S3_min_mmHg":          result.v_S3min,
            "t_SBP_S3_min":             result.t_S3e,
            "SBP_overshoot_mmHg":       result.v_ov,
            "t_SBP_overshoot":          result.t_ov,
            "t_SBP_back_to_baseline":   result.t_prt_end,
            "HR_max_bpm":               result.hr_max_v,
            "t_HR_max":                 result.hr_max_t,
            "HR_min_bpm":               result.hr_min_v,
            "t_HR_min":                 result.hr_min_t,
            "A_mmHg":                   result.A,
            "B_mmHg":                   result.B,
            "PRT_s":                    result.PRT,
            "VR":                       result.VR,
            "BRSa_mmHg_per_s":          result.BRSa,
        }

        sections = [
            ("Phase Boundaries (time in seconds)", [
                ("Baseline start",               "t_baseline_start",         "s",     "0.0"),
                ("Baseline end",                 "t_baseline_end",           "s",     "0.0"),
                ("PI start  (S1 start)",         "t_S1_start",               "s",     "0.0"),
                ("PI end  (S1 end)",             "t_S1_end",                 "s",     "0.0"),
                ("PII early end  (IIe nadir)",   "t_S2early_end",            "s",     "0.0"),
                ("PII late end  /  PIII start",  "t_S2late_end_PIII_start",  "s",     "0.0"),
                ("PIII end  /  PIV start",       "t_S3_end_PIV_start",       "s",     "0.0"),
                ("PIV end",                      "t_S4_end",                 "s",     "0.0"),
            ]),
            ("Signal Points", [
                ("Baseline SBP",               "SBP_baseline_mmHg",     "mmHg", "0.0"),
                ("IIe nadir SBP",              "SBP_IIe_nadir_mmHg",    "mmHg", "0.0"),
                ("S2late max SBP",             "SBP_S2late_max_mmHg",   "mmHg", "0.0"),
                ("S3 min SBP",                 "SBP_S3_min_mmHg",       "mmHg", "0.0"),
                ("S3 min time",                "t_SBP_S3_min",          "s",    "0.0"),
                ("SBP overshoot",              "SBP_overshoot_mmHg",    "mmHg", "0.0"),
                ("SBP overshoot time",         "t_SBP_overshoot",       "s",    "0.0"),
                ("SBP back to baseline time",  "t_SBP_back_to_baseline", "s",   "0.0"),
                ("HR max",                     "HR_max_bpm",            "bpm",  "0.0"),
                ("HR max time",                "t_HR_max",              "s",    "0.0"),
                ("HR min",                     "HR_min_bpm",            "bpm",  "0.0"),
                ("HR min time",                "t_HR_min",              "s",    "0.0"),
            ]),
            ("Derived Parameters  —  Novak 2011 / Mayo Clinic", [
                ("A  =  Baseline SBP − IIe nadir SBP",   "A_mmHg",          "mmHg",   "0.0"),
                ("B  =  S2late max SBP − S3 min SBP",    "B_mmHg",          "mmHg",   "0.0"),
                ("PRT  =  S3 nadir → SBP returns to BL", "PRT_s",           "s",      "0.00"),
                ("VR  =  HR max / HR min",               "VR",              "",       "0.00"),
                ("BRSa  =  (A + B×0.75) / PRT",         "BRSa_mmHg_per_s", "mmHg/s", "0.00"),
            ]),
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valsalva Analysis"

        def _w(row, col, val, font=None, fill=None, fmt=None, align="left"):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = font or val_font
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = _BRD
            if fill:
                c.fill = fill
            if fmt and val not in (None, "N/A"):
                c.number_format = fmt
            return c

        # Title
        ws.merge_cells("A1:D1")
        tc           = ws["A1"]
        tc.value     = "Valsalva Analysis"
        tc.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border    = _BRD
        ws.row_dimensions[1].height = 24

        for ci, h in enumerate(["Parameter", "Value", "Unit", "Note"], 1):
            _w(2, ci, h, font=hdr_font, fill=hdr_fill, align="center")
        ws.row_dimensions[2].height = 18

        row = 3
        for sect_title, items in sections:
            ws.merge_cells(f"A{row}:D{row}")
            _w(row, 1, sect_title, font=sect_font, fill=sect_fill)
            for ci in range(2, 5):
                ws.cell(row=row, column=ci).fill   = sect_fill
                ws.cell(row=row, column=ci).border = _BRD
            ws.row_dimensions[row].height = 16
            row += 1
            for label, key, unit, fmt in items:
                val  = data.get(key)
                note = ""
                fill = None
                if val is None:
                    disp, fill = "N/A", warn_fill
                else:
                    disp = round(float(val), 3)
                    if key == "A_mmHg" and float(val) < 0:
                        note = "IIe nadir above baseline — weak/absent strain response"
                        fill = warn_fill
                _w(row, 1, label, fill=fill)
                _w(row, 2, disp,  fill=fill, fmt=fmt, align="right")
                _w(row, 3, unit,  fill=fill, align="center")
                _w(row, 4, note,  fill=fill)
                ws.row_dimensions[row].height = 15
                row += 1

        for col_letter, width in zip("ABCD", [42, 12, 10, 52]):
            ws.column_dimensions[col_letter].width = width

        wb.save(out_path)
        log.info("Valsalva Excel saved: %s", out_path)
        return out_path

    def embed_images_valsalva(
        self,
        xlsx_path: Path,
        png_full: Path,
        png_zoom: Optional[Path],
    ) -> None:
        """Embed full and zoomed PNG plots into an existing Valsalva workbook."""
        self._embed_images(xlsx_path, [(png_full, "F1"), (png_zoom, "X1") if png_zoom else None])

    # ── Deep Breathing ───────────────────────────────────────────────────────

    def export_deep_breathing(
        self, result: DeepBreathingResult, output_dir: Path, mode: str = "auto"
    ) -> Path:
        """Write Deep Breathing RSA results to a timestamped, mode-tagged xlsx file."""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = output_dir / f"{output_dir.parent.name}_deep_breathing_results_{stamp}_{mode}.xlsx"

        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        ctr       = Alignment(horizontal="center", vertical="center")
        bold_blk  = Font(name="Arial", bold=True, size=11)
        norm      = Font(name="Arial", size=11)
        fill_green = PatternFill("solid", fgColor="E8F5E9")
        fill_grey  = PatternFill("solid", fgColor="F5F5F5")
        fill_sel   = PatternFill("solid", fgColor="F1F8E9")
        fill_alt   = PatternFill("solid", fgColor="EEF2F7")
        red_font   = Font(name="Arial", size=11, color="8B0000")
        blue_font  = Font(name="Arial", size=11, color="1A237E")
        red_bold   = Font(name="Arial", bold=True, size=11, color="8B0000")
        blue_bold  = Font(name="Arial", bold=True, size=11, color="1A237E")
        grey_bold  = Font(name="Arial", bold=True, size=11, color="555555")
        grn_bold   = Font(name="Arial", bold=True, size=11, color="1B5E20")
        chk_font   = Font(name="Arial", bold=True, size=11, color="2E7D32")
        thin       = Side(style="thin", color="CCCCCC")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deep Breathing RSA"

        def _wc(row, col, val, font=None, fill=None, align=None, brd=None):
            c = ws.cell(row=row, column=col, value=val)
            if font:  c.font      = font
            if fill:  c.fill      = fill
            if align: c.alignment = align
            if brd:   c.border    = brd
            return c

        ws.merge_cells("A1:G1")
        _wc(1, 1, "Deep Breathing — RSA Analysis",
            font=Font(name="Arial", bold=True, size=13),
            align=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[1].height = 22

        headers = ["#", "HR max (bpm)", "t max (s)", "HR min (bpm)", "t min (s)", "ΔHR (bpm)", "Top 6"]
        for ci, h in enumerate(headers, 1):
            _wc(2, ci, h, font=hdr_font, fill=hdr_fill, align=ctr, brd=border)
        ws.row_dimensions[2].height = 18

        for ri, c in enumerate(result.cycles):
            row = ri + 3
            selected = c.cycle in result.top6
            bg = fill_sel if selected else (fill_alt if ri % 2 else None)
            _wc(row, 1, c.cycle,              font=bold_blk if selected else norm, fill=bg, align=ctr, brd=border)
            _wc(row, 2, round(c.max_v, 1),   font=red_font,  fill=bg, align=ctr, brd=border)
            _wc(row, 3, round(c.max_t, 1),   font=norm,       fill=bg, align=ctr, brd=border)
            _wc(row, 4, round(c.min_v, 1),   font=blue_font,  fill=bg, align=ctr, brd=border)
            _wc(row, 5, round(c.min_t, 1),   font=norm,       fill=bg, align=ctr, brd=border)
            _wc(row, 6, round(c.rsa, 1),     font=norm,       fill=bg, align=ctr, brd=border)
            _wc(row, 7, "✓" if selected else "", font=chk_font if selected else norm,
                fill=bg, align=ctr, brd=border)

        n_data = len(result.cycles)

        mr_all = n_data + 3
        _wc(mr_all, 1, f"Mean (n={len(result.valid_cycles)})", font=bold_blk, fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 2, round(result.mean_max_all, 1), font=red_bold,  fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 3, "",                             font=norm,      fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 4, round(result.mean_min_all, 1), font=blue_bold, fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 5, "",                             font=norm,      fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 6, round(result.avg_rsa_all, 1),  font=grey_bold, fill=fill_grey, align=ctr, brd=border)
        _wc(mr_all, 7, "",                             font=norm,      fill=fill_grey, align=ctr, brd=border)

        mr_top = n_data + 4
        _wc(mr_top, 1, f"Mean (top {result.n_sel})", font=bold_blk, fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 2, round(result.mean_max_top, 1), font=red_bold,  fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 3, "",                             font=norm,      fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 4, round(result.mean_min_top, 1), font=blue_bold, fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 5, "",                             font=norm,      fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 6, round(result.avg_rsa_top6, 1), font=grn_bold,  fill=fill_green, align=ctr, brd=border)
        _wc(mr_top, 7, "✓",                            font=chk_font,  fill=fill_green, align=ctr, brd=border)

        col_min = [6, 10, 10, 12, 10, 10, 7]
        for ci in range(1, 8):
            max_len = col_min[ci - 1]
            for row_cells in ws.iter_rows(min_col=ci, max_col=ci):
                for cell in row_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(ci)].width = max_len

        wb.save(out_path)
        log.info("Deep Breathing Excel saved: %s", out_path)
        return out_path

    def embed_image_deep_breathing(self, xlsx_path: Path, png_path: Path) -> None:
        """Embed HR plot PNG into an existing Deep Breathing workbook."""
        self._embed_images(xlsx_path, [(png_path, "I1")])

    # ── shared helper ─────────────────────────────────────────────────────────

    def _embed_images(
        self,
        xlsx_path: Path,
        entries: list[Optional[tuple[Path, str]]],
    ) -> None:
        """Re-open workbook and embed PNGs at given anchors."""
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage

        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            for entry in entries:
                if entry is None:
                    continue
                png_path, anchor = entry
                if not Path(png_path).exists():
                    continue
                with PILImage.open(str(png_path)) as im:
                    nw, nh = im.size
                xl_img = XLImage(str(png_path))
                xl_img.width  = int(nw * 0.8)
                xl_img.height = int(nh * 0.8)
                ws.add_image(xl_img, anchor)
            wb.save(xlsx_path)
            log.info("Images embedded in %s", xlsx_path)
        except Exception as exc:
            log.error("Failed to embed images in %s: %s", xlsx_path, exc)
