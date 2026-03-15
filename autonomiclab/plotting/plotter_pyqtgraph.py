"""PyQtGraph plotting for interactive zoom/pan"""

import pyqtgraph as pg
import numpy as np


class GATPlotterPyQtGraph:
    """PyQtGraph-based plotter with mouse zoom/pan"""
    
    @staticmethod
    def _style_plot(plot):
        """Add box frame"""
        plot.showGrid(x=True, y=True, alpha=0.3)
        plot.getAxis('left').setPen(pg.mkPen(color='k', width=1))
        plot.getAxis('bottom').setPen(pg.mkPen(color='k', width=1))
        plot.getAxis('top').setPen(pg.mkPen(color='k', width=1))
        plot.getAxis('right').setPen(pg.mkPen(color='k', width=1))

    @staticmethod
    def _add_hr_variants(plot_widget, plot, dataset, t_start, t_end):
        """Add HR AP as line and HR ECG (RR-int) as small triangle markers"""
        # HR AP as line
        hr_ap_data = dataset.get('HR AP', {})
        if hr_ap_data:
            times = np.array(hr_ap_data.get('times', []))
            values = np.array(hr_ap_data.get('values', []))
            mask = (times >= t_start) & (times <= t_end)
            if np.any(mask):
                pen = pg.mkPen(color='#FF8C00', width=1.5)
                curve = plot.plot(times[mask], values[mask], pen=pen)
                plot_widget._plot_curves[id(plot)].append(curve)
                print("DEBUG: Added HR AP curve")

        # HR ECG (RR-int) as small triangle markers
        hr_ecg_data = dataset.get('HR ECG (RR-int)', {})
        if hr_ecg_data:
            times = np.array(hr_ecg_data.get('times', []))
            values = np.array(hr_ecg_data.get('values', []))
            mask = (times >= t_start) & (times <= t_end)
            if np.any(mask):
                curve = plot.plot(
                    times[mask], values[mask],
                    pen=None,
                    symbol='o', symbolSize=5,
                    symbolPen=pg.mkPen(color='#006400', width=1.5),
                    symbolBrush=pg.mkBrush(None),
                )
                # Add clean legend entry without symbol dot
                legend = plot.legend
                if legend is not None:
                    sample = pg.PlotDataItem(
                        pen=pg.mkPen(color='#006400', width=1.5))
                    legend.addItem(sample, 'HR ECG (RR-int)')
                plot_widget._plot_curves[id(plot)].append(curve)
                print("DEBUG: Added HR ECG (RR-int) markers")

    @staticmethod
    def create_overview_plot(plot_widget, dataset, markers):
        """Overview: BP (top), HR (middle), PAirway + Resp (bottom) — full recording"""
        print("DEBUG: create_overview_plot called")
        plot_widget.clear()

        if not hasattr(plot_widget, '_plot_curves'):
            plot_widget._plot_curves = {}

        # Determine full time range from HR signal
        hr_data = dataset.get('HR', {})
        if not hr_data:
            print("DEBUG: No HR data for overview")
            return

        hr_times = np.array(hr_data.get('times', []))
        hr_values = np.array(hr_data.get('values', []))
        if not len(hr_times):
            return

        t_start = float(hr_times[0])
        t_end   = float(hr_times[-1])

        plot1 = plot_widget.addPlot(row=0, col=0, title="Blood Pressure")
        plot2 = plot_widget.addPlot(row=1, col=0, title="Heart Rate")
        plot3 = plot_widget.addPlot(row=2, col=0, title="Airway / Respiratory")

        plot_widget._plot_curves[id(plot1)] = []
        plot_widget._plot_curves[id(plot2)] = []
        plot_widget._plot_curves[id(plot3)] = []

        plot2.setXLink(plot1)
        plot3.setXLink(plot1)

        # ── SUBPLOT 1: Blood Pressure ────────────────────────────────────────
        plot1.addLegend(offset=(10, 10))
        for sig, (color, width) in {
            'reBAP': ('#C0C0C0', 1),
            'reSYS': ('#FF0000', 2),
            'reDIA': ('#00AA00', 2),
            'reMAP': ('#0000FF', 2),
        }.items():
            d = dataset.get(sig, {})
            if d:
                t = np.array(d.get('times', []))
                v = np.array(d.get('values', []))
                if len(t):
                    curve = plot1.plot(t, v, pen=pg.mkPen(color=color, width=width), name=sig)
                    plot_widget._plot_curves[id(plot1)].append(curve)
        plot1.setLabel('left', 'BP (mmHg)')
        GATPlotterPyQtGraph._style_plot(plot1)

        # Phase marker lines on BP
        for marker in markers:
            plot1.addItem(pg.InfiniteLine(
                pos=marker['time'], angle=90,
                pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ── SUBPLOT 2: Heart Rate ────────────────────────────────────────────
        plot2.addLegend(offset=(10, 10))
        curve = plot2.plot(hr_times, hr_values,
                           pen=pg.mkPen(color='#8B0000', width=2), name='HR (AP)')
        plot_widget._plot_curves[id(plot2)].append(curve)
        GATPlotterPyQtGraph._add_hr_variants(plot_widget, plot2, dataset, t_start, t_end)
        plot2.setLabel('left', 'HR (bpm)')
        GATPlotterPyQtGraph._style_plot(plot2)

        for marker in markers:
            plot2.addItem(pg.InfiniteLine(
                pos=marker['time'], angle=90,
                pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ── SUBPLOT 3: PAirway + Resp Wave ───────────────────────────────────
        plot3.addLegend(offset=(10, 10))
        for sig, color, label in [
            ('PAirway',   '#0078d4', 'PAirway'),
            ('Resp Wave', '#00008B', 'Resp'),
        ]:
            d = dataset.get(sig, {})
            if d:
                t = np.array(d.get('times', []))
                v = np.array(d.get('values', []))
                if len(t):
                    curve = plot3.plot(t, v,
                                       pen=pg.mkPen(color=color, width=1.5), name=label)
                    plot_widget._plot_curves[id(plot3)].append(curve)
        plot3.setLabel('left', 'Amplitude')
        plot3.setLabel('bottom', 'Time (s)')
        GATPlotterPyQtGraph._style_plot(plot3)

        for marker in markers:
            plot3.addItem(pg.InfiniteLine(
                pos=marker['time'], angle=90,
                pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        plot1.setXRange(t_start, t_end)
        print("DEBUG: create_overview_plot complete")

    @staticmethod
    def _valsalva_analysis(plot_bp, plot_hr, plot_pa, dataset, markers, output_dir=None):
        """
        Novak 2011 / Mayo Clinic — 12 measurement points per reference figure.

        Phases (signal-derived):
          Baseline      [t_S1s−45, t_S1s−15]          (#1–#2)
          Anticipatory  [t_S1s−15, t_S1s]
          S1/PI         [t_S1s,    t_S1e]  PAirway↑0.5 → SBP local max  (#3–#4)
          S2early/IIe   [t_S1e,    t_S2es] → IIe nadir                   (#5)
          S2late/III    [t_S2es,   t_S3s]  → PAirway↓0.5                (#6–#7)
          S3/PIII       [t_S3s,    t_S3e]  → SBP local min post-release  (#8)
          S4/PIV        [t_S3e,    t_S4e]  → t(HR_max)+30s               (#9–#10)
          HR subplot:   HR max (#11), HR min (#12), VR bracket
        """
        import pyqtgraph as pg
        import numpy as np

        # ── Colour palette ───────────────────────────────────────────────────
        # Phase fill colours matching Novak 2011 illustration
        FILL = {
            'baseline':     (232, 244, 232, 130),   # green   – Baseline
            'anticipatory': (255, 255, 255,   0),   # transparent – Anticipatory (no fill)
            'S1':           (255, 243, 224, 120),   # orange  – PI
            'S2early':      (255, 235, 238, 120),   # pink    – PII early
            'S2late':       (255, 235, 238, 120),   # pink    – PII late
            'S3':           (237, 231, 246, 110),   # purple  – PIII
            'S4':           (224, 247, 250, 110),   # teal    – PIV
        }

        # Label / dot colours
        BL   = '#2e7d32'   # baseline green
        PI   = '#e65100'   # S1 orange
        PII  = '#c62828'   # S2 red
        PIII = '#283593'   # S3 dark blue
        PIV  = '#006064'   # S4 teal

        DashLine  = pg.QtCore.Qt.PenStyle.DashLine
        SolidLine = pg.QtCore.Qt.PenStyle.SolidLine

        # ── Helper draw functions ────────────────────────────────────────────
        def shade(plot, ta, tb, rgba):
            if ta is None or tb is None or tb <= ta: return
            r,g,b,a = rgba
            item = pg.LinearRegionItem(
                values=(ta, tb), orientation='vertical',
                brush=pg.mkBrush(r,g,b,a), pen=pg.mkPen(None), movable=False)
            item.setZValue(-10)
            plot.addItem(item)

        def vline(plot, t, color, style=DashLine, width=1.5):
            if t is None: return
            plot.addItem(pg.InfiniteLine(
                pos=t, angle=90,
                pen=pg.mkPen(color=color, width=width, style=style)))

        def hline_seg(plot, t1, t2, y, color, style=DashLine, width=1.5):
            if t1 is None or t2 is None: return
            plot.addItem(pg.PlotDataItem(
                x=[t1, t2], y=[y, y],
                pen=pg.mkPen(color=color, width=width, style=style)))

        def vline_seg(plot, t, y1, y2, color, style=SolidLine, width=2):
            if t is None or y1 is None or y2 is None: return
            plot.addItem(pg.PlotDataItem(
                x=[t, t], y=[y1, y2],
                pen=pg.mkPen(color=color, width=width, style=style)))

        def dot(plot, t, v, color, size=8):
            if t is None or v is None: return
            item = pg.ScatterPlotItem(
                x=[t], y=[v], size=size, symbol='o',
                pen=pg.mkPen(color, width=1.5),
                brush=pg.mkBrush(color))
            plot.addItem(item)
            # Prevent scatter item from appearing in legend
            if plot.legend is not None:
                try:
                    plot.legend.removeItem(item)
                except Exception:
                    pass

        def lbl(plot, t, v, txt, color, anchor=(0.5, 1.0), dy=0):
            if t is None or v is None: return
            item = pg.TextItem(txt, color=color, anchor=anchor)
            item.setPos(t, v + dy)
            plot.addItem(item)

        # ── Load signals ─────────────────────────────────────────────────────
        def _s(sig):
            d = dataset.get(sig, {})
            return np.array(d.get('times',[])), np.array(d.get('values',[]))

        t_pa, v_pa   = _s('PAirway')
        t_sys, v_sys = _s('reSYS')
        t_hr,  v_hr  = _s('HR')

        # ── Signal maths helpers ─────────────────────────────────────────────
        def mean_sys(ta, tb):
            m = (t_sys>=ta)&(t_sys<=tb)
            return float(np.mean(v_sys[m])) if np.any(m) else None

        def global_min_sys(ta, tb):
            m = (t_sys>=ta)&(t_sys<=tb)
            if not np.any(m): return None, None
            i = np.argmin(v_sys[m]); ts=t_sys[m]; vs=v_sys[m]
            return float(ts[i]), float(vs[i])

        def global_max_sys(ta, tb):
            m = (t_sys>=ta)&(t_sys<=tb)
            if not np.any(m): return None, None
            i = np.argmax(v_sys[m]); ts=t_sys[m]; vs=v_sys[m]
            return float(ts[i]), float(vs[i])

        def first_local_max_sys(ta, tb):
            """First 3-point local max in [ta,tb]."""
            m = (t_sys>=ta)&(t_sys<=tb)
            if not np.any(m): return None, None
            vs, ts = v_sys[m], t_sys[m]
            for i in range(1, len(vs)-1):
                if vs[i] > vs[i-1] and vs[i] > vs[i+1]:
                    return float(ts[i]), float(vs[i])
            return global_max_sys(ta, tb)   # fallback

        def first_local_min_sys(ta, tb):
            """First 3-point local min in [ta,tb]."""
            m = (t_sys>=ta)&(t_sys<=tb)
            if not np.any(m): return None, None
            vs, ts = v_sys[m], t_sys[m]
            for i in range(1, len(vs)-1):
                if vs[i] < vs[i-1] and vs[i] < vs[i+1]:
                    return float(ts[i]), float(vs[i])
            return global_min_sys(ta, tb)   # fallback

        THRESH = 0.5  # mmHg PAirway crossing threshold

        def pa_cross(t_from, direction='up', verify_level=5.0, verify_dur=1.0):
            """
            Find first PAirway crossing of THRESH after t_from.
            For 'up': verifies signal reaches verify_level within verify_dur s
            (avoids false triggers on brief spikes).
            Returns interpolated crossing time.
            """
            m = t_pa > t_from
            if not np.any(m): return None
            ts, vs = t_pa[m], v_pa[m]
            if direction == 'up':
                cands = np.where((vs[:-1] < THRESH) & (vs[1:] >= THRESH))[0]
            else:
                cands = np.where((vs[:-1] >= THRESH) & (vs[1:] < THRESH))[0]

            for ci in cands:
                t1, v1 = float(ts[ci]),   float(vs[ci])
                t2, v2 = float(ts[ci+1]), float(vs[ci+1])
                frac = (THRESH-v1)/(v2-v1) if (v2-v1) != 0 else 0.0
                t_cross = t1 + frac*(t2-t1)

                if direction == 'up':
                    # Must sustain above verify_level within verify_dur seconds
                    win = (ts >= t_cross) & (ts <= t_cross + verify_dur)
                    if np.any(win) and np.max(vs[win]) >= verify_level:
                        return t_cross
                else:
                    # Downward: accept first crossing
                    return t_cross

            # Fallback: return first crossing without verify
            if len(cands) > 0:
                ci = cands[0]
                t1,v1 = float(ts[ci]),float(vs[ci])
                t2,v2 = float(ts[ci+1]),float(vs[ci+1])
                frac = (THRESH-v1)/(v2-v1) if (v2-v1) != 0 else 0.0
                return t1 + frac*(t2-t1)
            return None

        def find_marker(substr):
            for mk in markers:
                if substr.lower() in mk.get('label','').lower():
                    return mk['time']
            return None

        # ════════════════════════════════════════════════════════════════════
        # PHASE BOUNDARY DETECTION
        # ════════════════════════════════════════════════════════════════════
        t_vm1    = find_marker('VM1')
        t_anchor = t_vm1 if t_vm1 is not None else (float(t_pa[0]) if len(t_pa) else 0.0)

        # --- #3  S1 start = PAirway ↑ 0.5 (sustained) -----------------------
        t_S1s = pa_cross(t_anchor, 'up', verify_level=5.0, verify_dur=1.0)

        # Baseline window: [t_S1s−45, t_S1s−15]
        t_bl_s = (t_S1s - 45.0) if t_S1s else None   # #1
        t_bl_e = (t_S1s - 15.0) if t_S1s else None   # #2

        # --- #7  S3 start = PAirway ↓ 0.5 -----------------------------------
        t_S3s = pa_cross(t_S1s or t_anchor, 'down')

        # --- #4  S1 end = first local SBP max in [t_S1s, t_S3s] -------------
        t_S1e, _ = first_local_max_sys(t_S1s, t_S3s) if t_S1s and t_S3s else (None, None)

        # --- #5  IIe nadir = global SBP min in [t_S1e, t_S3s] ---------------
        t_S2es, v_nadir = global_min_sys(t_S1e, t_S3s) if t_S1e and t_S3s else (None, None)

        # --- #6  S2late max = global SBP max in [t_S2es, t_S3s] -------------
        t_S2lmax, v_S2lmax = global_max_sys(t_S2es, t_S3s) if t_S2es and t_S3s else (None, None)

        # --- #8  S3 end = global SBP min post-release in [t_S3s, t_S3s+20s]
        # Use global min — the SBP nadir after release is the deepest point,
        # not necessarily the first local minimum.
        t_S3e, v_S3min = global_min_sys(t_S3s, t_S3s+20.0) if t_S3s else (None, None)

        # --- #11 HR max in [t_S3s, t_S3s+8s]
        # Novak 2011 (J Vis Exp 53:e2502): HR peaks shortly after release (t_S3s).
        # Window capped at t_S3s+8s (cf. Denq 1998: ≤15s post-release).
        hr_max_t, hr_max_v = None, None
        t_hr_max_end = (t_S3s + 8.0) if t_S3s else None
        if t_S3s and t_hr_max_end and len(t_hr):
            m = (t_hr >= t_S3s) & (t_hr <= t_hr_max_end)
            if np.any(m):
                i = np.argmax(v_hr[m])
                hr_max_t, hr_max_v = float(t_hr[m][i]), float(v_hr[m][i])

        # S4 end = t(HR max) + 30s
        t_S4e = (hr_max_t + 30.0) if hr_max_t else None

        # --- #12 HR min in [t(HR max), t(HR max)+30s] -----------------------
        hr_min_t, hr_min_v = None, None
        if hr_max_t and t_S4e and len(t_hr):
            m = (t_hr >= hr_max_t) & (t_hr <= t_S4e)
            if np.any(m):
                i = np.argmin(v_hr[m])
                hr_min_t, hr_min_v = float(t_hr[m][i]), float(v_hr[m][i])

        # Baseline SBP = mean reSYS in [t_bl_s, t_bl_e]
        avg_sbp = mean_sys(t_bl_s, t_bl_e) if t_bl_s else None

        print(f"DEBUG BL=[{t_bl_s},{t_bl_e}] avg_sbp={avg_sbp}")
        print(f"DEBUG S1=[{t_S1s},{t_S1e}] S2early=[{t_S1e},{t_S2es}] v_nadir={v_nadir}")
        print(f"DEBUG S2late_max@{t_S2lmax}={v_S2lmax}  S2late=[{t_S2es},{t_S3s}]")
        print(f"DEBUG S3=[{t_S3s},{t_S3e}] v_S3min={v_S3min}  S4=[{t_S3e},{t_S4e}]")
        print(f"DEBUG HR_max={hr_max_v}@{hr_max_t}  HR_min={hr_min_v}@{hr_min_t}")

        # ════════════════════════════════════════════════════════════════════
        # DRAW PHASE BOXES on all 3 subplots
        # ════════════════════════════════════════════════════════════════════
        for p in (plot_bp, plot_hr, plot_pa):
            shade(p, t_bl_s,  t_bl_e,  FILL['baseline'])
            shade(p, t_bl_e,  t_S1s,   FILL['anticipatory'])
            shade(p, t_S1s,   t_S1e,   FILL['S1'])
            shade(p, t_S1e,   t_S2es,  FILL['S2early'])
            shade(p, t_S2es,  t_S3s,   FILL['S2late'])
            shade(p, t_S3s,   t_S3e,   FILL['S3'])
            shade(p, t_S3e,   t_S4e,   FILL['S4'])

        # ════════════════════════════════════════════════════════════════════
        # BP SUBPLOT MEASUREMENTS
        # ════════════════════════════════════════════════════════════════════

        # Baseline SBP dashed horizontal line
        if avg_sbp is not None:
            plot_bp.addItem(pg.InfiniteLine(
                pos=avg_sbp, angle=0,
                pen=pg.mkPen(color=BL, width=1.5, style=DashLine)))

        # Points #1 and #2: baseline window boundary lines
        vline(plot_bp, t_bl_s, BL, DashLine)
        vline(plot_bp, t_bl_e, BL, DashLine)

        # Point #3: S1 start (PAirway ↑) — vertical line on all subplots
        for p in (plot_bp, plot_hr, plot_pa):
            vline(p, t_S1s, PI, DashLine, width=2)

        # Point #4: SBP local max in S1 — dot at S1/S2early boundary
        t_S1e_t, v_S1e_v = first_local_max_sys(t_S1s, t_S3s) if t_S1s and t_S3s else (None,None)
        dot(plot_bp, t_S1e_t, v_S1e_v, PI)

        # Point #5: IIe nadir + measurement A
        A = None
        if t_S2es and v_nadir is not None:
            dot(plot_bp, t_S2es, v_nadir, PII)
            # Vertical dashed arrow A: nadir → baseline
            if avg_sbp:
                vline_seg(plot_bp, t_S2es, v_nadir, avg_sbp, PII, DashLine)
                A = avg_sbp - v_nadir
                lbl(plot_bp, t_S2es - 1.5, (v_nadir + avg_sbp)/2,
                    f'A={A:.0f}', PII, anchor=(1.0, 0.5))
                print(f"DEBUG A={A:.1f} mmHg")

        # Point #6: SBP max in S2late
        B = None
        if t_S2lmax and v_S2lmax:
            dot(plot_bp, t_S2lmax, v_S2lmax, PI)

        # Point #7: PAirway↓ boundary line — already shown as S3 shade edge
        for p in (plot_bp, plot_hr, plot_pa):
            vline(p, t_S3s, PIII, DashLine, width=2)

        # Point #8: SBP local min in S3 + measurement B
        if t_S3e and v_S3min is not None:
            dot(plot_bp, t_S3e, v_S3min, PIII)
            if t_S2lmax and v_S2lmax:
                # Place B bracket just before the PIII boundary (t_S3s)
                t_b = (t_S3s - 2.0) if t_S3s else (t_S2lmax + t_S3e) / 2
                hline_seg(plot_bp, t_S2lmax, t_b, v_S2lmax, PI,   DashLine)
                hline_seg(plot_bp, t_b,      t_b, v_S3min,  PIII, DashLine)
                # Extend S3min hline leftward to bracket
                hline_seg(plot_bp, t_S2lmax, t_b, v_S3min,  PIII, DashLine)
                # Vertical B bracket
                vline_seg(plot_bp, t_b, v_S3min, v_S2lmax, PIII)
                B = v_S2lmax - v_S3min
                lbl(plot_bp, t_b - 0.5, (v_S2lmax + v_S3min) / 2,
                    f'B={B:.0f}', PIII, anchor=(1.0, 0.5))
                print(f"DEBUG B={B:.1f} mmHg")

        # Point #10: SBP overshoot = global SBP max in S4
        t_ov, v_ov = global_max_sys(t_S3e, t_S4e) if t_S3e and t_S4e else (None, None)
        if t_ov and v_ov:
            dot(plot_bp, t_ov, v_ov, PIV)
            lbl(plot_bp, t_ov, v_ov, 'SBP\nOvershoot', PIV, anchor=(0.5,1.0), dy=2)

        # Point #9: PRT = t(SBP returns to baseline) − t(SBP min in S3 = t_S3e)
        PRT = None
        if t_S3e and avg_sbp and len(t_sys):
            m = t_sys > t_S3e
            if np.any(m):
                ta, va = t_sys[m], v_sys[m]
                ci_arr = np.where(va >= avg_sbp)[0]
                if len(ci_arr):
                    ci = ci_arr[0]
                    if ci > 0:
                        t1,v1 = float(ta[ci-1]),float(va[ci-1])
                        t2,v2 = float(ta[ci]),  float(va[ci])
                        frac = (avg_sbp-v1)/(v2-v1) if (v2-v1)!=0 else 0.0
                        t_prt_end = t1 + frac*(t2-t1)
                    else:
                        t_prt_end = float(ta[0])
                    PRT = t_prt_end - t_S3e
                    print(f"DEBUG PRT={PRT:.2f}s")
                    dot(plot_bp, t_prt_end, avg_sbp, PIII)
                    hline_seg(plot_bp, t_S3e, t_prt_end, avg_sbp, PIII, SolidLine, 2)
                    lbl(plot_bp, (t_S3e+t_prt_end)/2, avg_sbp,
                        f'PRT={PRT:.1f}s', PIII, anchor=(0.5,1.0), dy=2)

        # ════════════════════════════════════════════════════════════════════
        # HR SUBPLOT MEASUREMENTS  (#11, #12, VR)
        # ════════════════════════════════════════════════════════════════════
        if hr_max_t and hr_max_v:
            dot(plot_hr, hr_max_t, hr_max_v, PIII)
            lbl(plot_hr, hr_max_t, hr_max_v, f'HR max\n{hr_max_v:.1f}',
                PIII, anchor=(0.5,1.0), dy=1)

        if hr_min_t and hr_min_v:
            dot(plot_hr, hr_min_t, hr_min_v, PIV)
            lbl(plot_hr, hr_min_t, hr_min_v, f'HR min\n{hr_min_v:.1f}',
                PIV, anchor=(0.5,0.0), dy=-1)

        if hr_max_t and hr_min_t and t_S4e:
            # Dashed horizontals at HR max and HR min level across S4
            hline_seg(plot_hr, hr_max_t, t_S4e, hr_max_v, PIII, DashLine)
            hline_seg(plot_hr, hr_min_t, t_S4e, hr_min_v, PIV,  DashLine)
            # Vertical VR bracket at right edge of S4
            vline_seg(plot_hr, t_S4e, hr_min_v, hr_max_v, PIV)
            if hr_min_v and hr_min_v > 0:
                VR = hr_max_v / hr_min_v
                lbl(plot_hr, t_S4e+1, (hr_max_v+hr_min_v)/2,
                    f'VR={VR:.2f}', PIV, anchor=(0.0,0.5))
                print(f"DEBUG VR={VR:.2f}")
                if A and B and PRT and PRT > 0:
                    BRSa = (A + B*0.75) / PRT
                    print(f"DEBUG BRSa={BRSa:.2f} mmHg/s")

        # ── HR subplot y-range: +10 bpm headroom above HR max ────────────────
        if hr_max_v and len(t_hr):
            t_zoom_s = (t_S1s - 60) if t_S1s else float(t_hr[0])
            t_zoom_e = t_S4e if t_S4e else float(t_hr[-1])
            m = (t_hr >= t_zoom_s) & (t_hr <= t_zoom_e)
            y_min = float(np.min(v_hr[m])) - 10 if np.any(m) else 40
            plot_hr.setYRange(y_min, hr_max_v + 10)

        # ── Excel export ──────────────────────────────────────────────────────
        VR_val   = (hr_max_v / hr_min_v) if (hr_max_v and hr_min_v and hr_min_v > 0) else None
        BRSa_val = ((A + B * 0.75) / PRT) if (A is not None and B and PRT and PRT > 0) else None
        t_prt_end_val = locals().get('t_prt_end', None)

        results = {
            't_baseline_start':          t_bl_s,
            't_baseline_end':            t_bl_e,
            't_S1_start':                t_S1s,
            't_S1_end':                  t_S1e,
            't_S2early_end':             t_S2es,
            't_S2late_end_PIII_start':   t_S3s,
            't_S3_end_PIV_start':        t_S3e,
            't_S4_end':                  t_S4e,
            'SBP_baseline_mmHg':         avg_sbp,
            'SBP_IIe_nadir_mmHg':        v_nadir,
            'SBP_S2late_max_mmHg':       v_S2lmax,
            'SBP_S3_min_mmHg':           v_S3min,
            't_SBP_S3_min':              t_S3e,
            'SBP_overshoot_mmHg':        v_ov,
            't_SBP_overshoot':           t_ov,
            't_SBP_back_to_baseline':    t_prt_end_val,
            'HR_max_bpm':                hr_max_v,
            't_HR_max':                  hr_max_t,
            'HR_min_bpm':                hr_min_v,
            't_HR_min':                  hr_min_t,
            'A_mmHg':                    A,
            'B_mmHg':                    B,
            'PRT_s':                     PRT,
            'VR':                        VR_val,
            'BRSa_mmHg_per_s':           BRSa_val,
        }

        try:
            try:
                import openpyxl
            except ImportError:
                import subprocess, sys
                subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                                       'openpyxl', '--quiet', '--break-system-packages'],
                                      stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import os

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Valsalva Analysis'

            hdr_fill  = PatternFill('solid', start_color='1F4E79')
            sect_fill = PatternFill('solid', start_color='D6E4F0')
            warn_fill = PatternFill('solid', start_color='FFF2CC')
            hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            sect_font = Font(name='Arial', bold=True, color='1F4E79', size=10)
            val_font  = Font(name='Arial', size=10)
            thin      = Side(style='thin', color='AAAAAA')
            brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _w(row, col, val, font=None, fill=None, fmt=None, align='left'):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = font or val_font
                c.alignment = Alignment(horizontal=align, vertical='center')
                c.border    = brd
                if fill: c.fill = fill
                if fmt and val not in (None, 'N/A'): c.number_format = fmt
                return c

            sections = [
                ('Phase Boundaries (time in seconds)', [
                    ('Baseline start',              't_baseline_start',         's', '0.0'),
                    ('Baseline end',                't_baseline_end',           's', '0.0'),
                    ('PI start  (S1 start)',         't_S1_start',               's', '0.0'),
                    ('PI end  (S1 end)',             't_S1_end',                 's', '0.0'),
                    ('PII early end  (IIe nadir)',   't_S2early_end',            's', '0.0'),
                    ('PII late end  /  PIII start',  't_S2late_end_PIII_start',  's', '0.0'),
                    ('PIII end  /  PIV start',       't_S3_end_PIV_start',       's', '0.0'),
                    ('PIV end',                      't_S4_end',                 's', '0.0'),
                ]),
                ('Signal Points', [
                    ('Baseline SBP',                'SBP_baseline_mmHg',     'mmHg', '0.0'),
                    ('IIe nadir SBP',               'SBP_IIe_nadir_mmHg',    'mmHg', '0.0'),
                    ('S2late max SBP',              'SBP_S2late_max_mmHg',   'mmHg', '0.0'),
                    ('S3 min SBP',                  'SBP_S3_min_mmHg',       'mmHg', '0.0'),
                    ('S3 min time',                 't_SBP_S3_min',          's',    '0.0'),
                    ('SBP overshoot',               'SBP_overshoot_mmHg',    'mmHg', '0.0'),
                    ('SBP overshoot time',          't_SBP_overshoot',       's',    '0.0'),
                    ('SBP back to baseline time',   't_SBP_back_to_baseline', 's',   '0.0'),
                    ('HR max',                      'HR_max_bpm',            'bpm',  '0.0'),
                    ('HR max time',                 't_HR_max',              's',    '0.0'),
                    ('HR min',                      'HR_min_bpm',            'bpm',  '0.0'),
                    ('HR min time',                 't_HR_min',              's',    '0.0'),
                ]),
                ('Derived Parameters  —  Novak 2011 / Mayo Clinic', [
                    ('A  =  Baseline SBP − IIe nadir SBP',   'A_mmHg',          'mmHg',   '0.0'),
                    ('B  =  S2late max SBP − S3 min SBP',    'B_mmHg',          'mmHg',   '0.0'),
                    ('PRT  =  S3 nadir → SBP returns to BL', 'PRT_s',           's',      '0.00'),
                    ('VR  =  HR max / HR min',               'VR',              '',       '0.00'),
                    ('BRSa  =  (A + B×0.75) / PRT',         'BRSa_mmHg_per_s', 'mmHg/s', '0.00'),
                ]),
            ]

            # Title row
            ws.merge_cells('A1:D1')
            tc           = ws['A1']
            tc.value     = 'Valsalva Analysis'
            tc.font      = Font(name='Arial', bold=True, size=13, color='1F4E79')
            tc.alignment = Alignment(horizontal='center', vertical='center')
            tc.border    = brd
            ws.row_dimensions[1].height = 24

            # Column headers
            for ci, h in enumerate(['Parameter', 'Value', 'Unit', 'Note'], 1):
                _w(2, ci, h, font=hdr_font, fill=hdr_fill, align='center')
            ws.row_dimensions[2].height = 18

            row = 3
            for sect_title, items in sections:
                ws.merge_cells(f'A{row}:D{row}')
                _w(row, 1, sect_title, font=sect_font, fill=sect_fill)
                for ci in range(2, 5):
                    ws.cell(row=row, column=ci).fill   = sect_fill
                    ws.cell(row=row, column=ci).border = brd
                ws.row_dimensions[row].height = 16
                row += 1
                for label, key, unit, fmt in items:
                    val  = results.get(key)
                    note = ''
                    fill = None
                    if val is None:
                        disp, fill = 'N/A', warn_fill
                    else:
                        disp = round(float(val), 3)
                        if key == 'A_mmHg' and float(val) < 0:
                            note = 'IIe nadir above baseline — weak/absent strain response'
                            fill = warn_fill
                    _w(row, 1, label, fill=fill)
                    _w(row, 2, disp,  fill=fill, fmt=fmt, align='right')
                    _w(row, 3, unit,  fill=fill, align='center')
                    _w(row, 4, note,  fill=fill)
                    ws.row_dimensions[row].height = 15
                    row += 1

            for col_letter, width in zip('ABCD', [42, 12, 10, 52]):
                ws.column_dimensions[col_letter].width = width

            # Save in output_dir (data folder) if provided, else beside script
            if output_dir and os.path.isdir(output_dir):
                _prefix = os.path.basename(os.path.normpath(output_dir))
                out_path = os.path.join(output_dir, f'{_prefix}_valsalva_results.xlsx')
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                out_path   = os.path.join(script_dir, 'valsalva_results.xlsx')
            wb.save(out_path)
            print(f"Excel gemt: {out_path}")

        except Exception as e:
            print(f"DEBUG Excel export failed: {e}")

        return t_S1s, t_S4e

    @staticmethod
    def create_valsalva_plot(plot_widget, dataset, markers, t_start=50, t_end=350, output_dir=None):
        """Valsalva: BP (top), HR + HR ECG (middle), PAirway (bottom)"""
        print("DEBUG: create_valsalva_plot called")
        plot_widget.clear()

        bp_signals = {
            'reBAP': ("#808080", 1),
            'reSYS': ('#FF0000', 2),
            'reDIA': ('#00AA00', 2),
            'reMAP': ('#0000FF', 2)
        }

        hr_data = dataset.get('HR', {})
        pairway_data = dataset.get('PAirway', {})

        if not hr_data or not pairway_data:
            print("DEBUG: Missing HR or PAirway data")
            return

        hr_times = np.array(hr_data.get('times', []))
        hr_values = np.array(hr_data.get('values', []))
        pairway_times = np.array(pairway_data.get('times', []))
        pairway_values = np.array(pairway_data.get('values', []))

        hr_mask = (hr_times >= t_start) & (hr_times <= t_end)
        pa_mask = (pairway_times >= t_start) & (pairway_times <= t_end)

        plot1 = plot_widget.addPlot(row=0, col=0, title="Blood Pressure (Multiple Traces)", rowspan=1)
        plot2 = plot_widget.addPlot(row=1, col=0, title="Heart Rate", rowspan=2)
        plot3 = plot_widget.addPlot(row=3, col=0, title="Airway Pressure", rowspan=1)

        if not hasattr(plot_widget, '_plot_curves'):
            plot_widget._plot_curves = {}
        plot_widget._plot_curves[id(plot1)] = []
        plot_widget._plot_curves[id(plot2)] = []
        plot_widget._plot_curves[id(plot3)] = []

        plot2.setXLink(plot1)
        plot3.setXLink(plot1)

        # ===== SUBPLOT 1: Blood Pressure =====
        plot1.addLegend(offset=(10, 10))

        for signal_name, (color, width) in bp_signals.items():
            bp_data = dataset.get(signal_name, {})
            if bp_data:
                bp_times = np.array(bp_data.get('times', []))
                bp_values = np.array(bp_data.get('values', []))
                bp_mask = (bp_times >= t_start) & (bp_times <= t_end)
                if np.any(bp_mask):
                    pen = pg.mkPen(color=color, width=width)
                    curve = plot1.plot(bp_times[bp_mask], bp_values[bp_mask], pen=pen, name=signal_name)
                    plot_widget._plot_curves[id(plot1)].append(curve)

        plot1.setLabel('left', 'BP (mmHg)')
        GATPlotterPyQtGraph._style_plot(plot1)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot1.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ===== SUBPLOT 2: HR (AP) only =====
        plot2.addLegend(offset=(10, 10))

        pen_hr = pg.mkPen(color='#8B0000', width=2.5)
        hr_curve = plot2.plot(hr_times[hr_mask], hr_values[hr_mask], pen=pen_hr, name='HR (AP)')
        plot_widget._plot_curves[id(plot2)].append(hr_curve)

        plot2.setLabel('left', 'HR (bpm)')
        GATPlotterPyQtGraph._style_plot(plot2)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot2.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ===== SUBPLOT 3: Airway Pressure =====
        plot3.addLegend(offset=(10, 10))

        pen_pa = pg.mkPen(color='#0078d4', width=2)
        pa_curve = plot3.plot(pairway_times[pa_mask], pairway_values[pa_mask], pen=pen_pa, name='PAirway')
        plot_widget._plot_curves[id(plot3)].append(pa_curve)

        plot3.setLabel('left', 'PAirway (mmHg)')
        plot3.setLabel('bottom', 'Time (s)')
        GATPlotterPyQtGraph._style_plot(plot3)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot3.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        plot1.setXRange(t_start, t_end)
        zoom = GATPlotterPyQtGraph._valsalva_analysis(plot1, plot2, plot3, dataset, markers, output_dir=output_dir)
        if zoom[0] is not None and zoom[1] is not None:
            plot1.setXRange(zoom[0] - 60, zoom[1] + 15)

        # ── Export plot as PNG and embed in Excel ─────────────────────────────
        try:
            import os
            _out_dir   = output_dir if (output_dir and os.path.isdir(output_dir)) \
                         else os.path.dirname(os.path.abspath(__file__))
            _prefix    = os.path.basename(os.path.normpath(_out_dir))
            png_path   = os.path.join(_out_dir, f'{_prefix}_valsalva_plot.png')
            xlsx_path  = os.path.join(_out_dir, f'{_prefix}_valsalva_results.xlsx')

            # Render full plot
            from pyqtgraph.exporters import ImageExporter
            exporter = ImageExporter(plot_widget.scene())
            exporter.export(png_path)
            print(f"PNG gemt: {png_path}")

            # Render zoomed plot [t_S1s-5, t_S4e+5]
            png_zoom_path = os.path.join(_out_dir, f'{_prefix}_valsalva_plot_zoom.png')
            if zoom[0] is not None and zoom[1] is not None:
                z_s, z_e = zoom[0] - 5, zoom[1] + 5
                p1 = plot_widget.getItem(0, 0)
                p2 = plot_widget.getItem(1, 0)
                p3 = plot_widget.getItem(3, 0)
                # Break x-link temporarily so setXRange takes effect independently
                for p in (p2, p3):
                    if p: p.setXLink(None)
                for p in (p1, p2, p3):
                    if p: p.setXRange(z_s, z_e, padding=0)
                pg.QtWidgets.QApplication.processEvents()
                exporter2 = ImageExporter(plot_widget.scene())
                exporter2.export(png_zoom_path)
                # Restore x-link
                for p in (p2, p3):
                    if p and p1: p.setXLink(p1)
                print(f"PNG zoom gemt: {png_zoom_path}")
            else:
                png_zoom_path = None

            # Embed PNG in existing Excel file
            if os.path.exists(xlsx_path) and os.path.exists(png_path):
                try:
                    import openpyxl
                except ImportError:
                    import subprocess, sys
                    subprocess.check_call(
                        [sys.executable, '-m', 'pip', 'install', 'openpyxl',
                         '--quiet', '--break-system-packages'],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    import openpyxl
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image as XLImage
                from PIL import Image as PILImage

                wb = load_workbook(xlsx_path)
                ws = wb.active

                def _embed(ws, png, anchor):
                    if not os.path.exists(png): return
                    with PILImage.open(png) as im:
                        nw, nh = im.size
                    tw = 1120
                    th = int(tw * nh / nw)
                    img = XLImage(png)
                    img.width  = tw
                    img.height = th
                    ws.add_image(img, anchor)

                _embed(ws, png_path, 'F1')
                if png_zoom_path and os.path.exists(png_zoom_path):
                    _embed(ws, png_zoom_path, 'X1')
                wb.save(xlsx_path)
                print(f"Billeder indlejret i Excel: {xlsx_path}")

        except Exception as e:
            print(f"DEBUG PNG/Excel image export failed: {e}")

        print("DEBUG: create_valsalva_plot complete")

    @staticmethod
    def create_stand_test_plot(plot_widget, dataset, markers, t_start=320, t_end=1020):
        """Stand Test: BP (top), HR + HR ECG (bottom)"""
        print("DEBUG: create_stand_test_plot called")
        plot_widget.clear()

        bp_signals = {
            'reBAP': ("#C4C4C4", 1),
            'reSYS': ('#FF0000', 2),
            'reDIA': ('#00AA00', 2),
            'reMAP': ('#0000FF', 2)
        }

        hr_data = dataset.get('HR', {})

        if not hr_data:
            print("DEBUG: Missing HR data")
            return

        hr_times = np.array(hr_data.get('times', []))
        hr_values = np.array(hr_data.get('values', []))
        hr_mask = (hr_times >= t_start) & (hr_times <= t_end)

        plot1 = plot_widget.addPlot(row=0, col=0, title="Blood Pressure (Multiple Traces)", rowspan=1)
        plot2 = plot_widget.addPlot(row=1, col=0, title="Heart Rate", rowspan=2)

        if not hasattr(plot_widget, '_plot_curves'):
            plot_widget._plot_curves = {}
        plot_widget._plot_curves[id(plot1)] = []
        plot_widget._plot_curves[id(plot2)] = []

        plot2.setXLink(plot1)

        # ===== SUBPLOT 1: Blood Pressure =====
        plot1.addLegend(offset=(10, 10))

        for signal_name, (color, width) in bp_signals.items():
            bp_data = dataset.get(signal_name, {})
            if bp_data:
                bp_times = np.array(bp_data.get('times', []))
                bp_values = np.array(bp_data.get('values', []))
                bp_mask = (bp_times >= t_start) & (bp_times <= t_end)
                if np.any(bp_mask):
                    pen = pg.mkPen(color=color, width=width)
                    curve = plot1.plot(bp_times[bp_mask], bp_values[bp_mask], pen=pen, name=signal_name)
                    plot_widget._plot_curves[id(plot1)].append(curve)

        plot1.setLabel('left', 'BP (mmHg)')
        GATPlotterPyQtGraph._style_plot(plot1)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot1.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ===== SUBPLOT 2: HR (AP) + HR ECG =====
        plot2.addLegend(offset=(10, 10))

        pen_hr = pg.mkPen(color='#8B0000', width=2)
        hr_curve = plot2.plot(hr_times[hr_mask], hr_values[hr_mask], pen=pen_hr, name='HR (AP)')
        plot_widget._plot_curves[id(plot2)].append(hr_curve)

        GATPlotterPyQtGraph._add_hr_variants(plot_widget, plot2, dataset, t_start, t_end)

        plot2.setLabel('left', 'HR (bpm)')
        plot2.setLabel('bottom', 'Time (s)')
        GATPlotterPyQtGraph._style_plot(plot2)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot2.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        plot1.setXRange(t_start, t_end)
        print("DEBUG: create_stand_test_plot complete")

    @staticmethod
    def create_deep_breathing_plot(plot_widget, dataset, markers, t_start=980, t_end=1450, output_dir=None):
        """Deep Breathing: HR + HR ECG (top), Resp Wave (bottom)"""
        print("DEBUG: create_deep_breathing_plot called")
        plot_widget.clear()

        hr_data = dataset.get('HR', {})

        if not hr_data:
            print("DEBUG: Missing HR data")
            return

        hr_times = np.array(hr_data.get('times', []))
        hr_values = np.array(hr_data.get('values', []))

        hr_mask = (hr_times >= t_start) & (hr_times <= t_end)

        plot1 = plot_widget.addPlot(row=0, col=0, title="Heart Rate (RSA)", rowspan=1)

        if not hasattr(plot_widget, '_plot_curves'):
            plot_widget._plot_curves = {}
        plot_widget._plot_curves[id(plot1)] = []

        # ===== SUBPLOT 1: HR (AP) only =====
        plot1.addLegend(offset=(10, 10))

        pen_hr = pg.mkPen(color='#8B0000', width=2)
        hr_curve = plot1.plot(hr_times[hr_mask], hr_values[hr_mask], pen=pen_hr, name='HR (AP)')
        plot_widget._plot_curves[id(plot1)].append(hr_curve)

        plot1.setLabel('left', 'HR (bpm)')
        GATPlotterPyQtGraph._style_plot(plot1)

        for marker in markers:
            if t_start <= marker['time'] <= t_end:
                plot1.addItem(pg.InfiniteLine(pos=marker['time'], angle=90,
                              pen=pg.mkPen('r', width=1, style=pg.QtCore.Qt.PenStyle.DashLine)))

        # ===== ROW 1: QTableWidget via GraphicsProxyWidget =====
        from pyqtgraph.Qt import QtWidgets, QtCore
        table_widget = QtWidgets.QTableWidget()
        table_widget.setColumnCount(7)
        table_widget.setHorizontalHeaderLabels(['#', 'HR max', 't max (s)', 'HR min', 't min (s)', 'ΔHR', 'Top 6'])
        table_widget.horizontalHeader().setStretchLastSection(True)
        table_widget.horizontalHeader().setSectionResizeMode(
            QtWidgets.QHeaderView.ResizeMode.Stretch)
        table_widget.verticalHeader().setVisible(False)
        table_widget.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        table_widget.setAlternatingRowColors(True)
        table_widget.setStyleSheet("""
            QTableWidget { font-size: 12px; gridline-color: #cccccc; }
            QHeaderView::section { background-color: #1F4E79; color: white;
                                   font-weight: bold; padding: 4px; }
            QTableWidget::item:alternate { background-color: #f0f4f8; }
        """)

        from pyqtgraph.Qt.QtWidgets import QGraphicsProxyWidget
        proxy_item = QGraphicsProxyWidget()
        proxy_item.setWidget(table_widget)
        plot_widget.ci.addItem(proxy_item, row=1, col=0)

        # 50/50 split
        plot_widget.ci.layout.setRowStretchFactor(0, 1)
        plot_widget.ci.layout.setRowStretchFactor(1, 1)

        # ===== DBM2–DBM3: Guided breathing phase box =====
        def find_marker(substr):
            for mk in markers:
                lbl = mk.get('label', '')
                if substr.lower() in lbl.lower():
                    return mk['time']
            return None

        print(f"DEBUG DB markers: {[(mk['time'], mk.get('label','')) for mk in markers]}")
        t_dbm2 = find_marker('DBM2')
        t_dbm3 = find_marker('DBM3')
        print(f"DEBUG t_dbm2={t_dbm2}  t_dbm3={t_dbm3}")

        if t_dbm2 and t_dbm3:
            FILL_DB = (224, 240, 255, 120)
            region = pg.LinearRegionItem(
                values=(t_dbm2, t_dbm3), orientation='vertical',
                brush=pg.mkBrush(*FILL_DB), pen=pg.mkPen(None), movable=False)
            region.setZValue(-10)
            plot1.addItem(region)
            for t, color in ((t_dbm2, '#1565C0'), (t_dbm3, '#1565C0')):
                plot1.addItem(pg.InfiniteLine(
                    pos=t, angle=90,
                    pen=pg.mkPen(color=color, width=2,
                                 style=pg.QtCore.Qt.PenStyle.DashLine)))
            lbl_item = pg.TextItem('Guided breathing\n5s insp / 5s exp',
                                   color='#1565C0', anchor=(0.5, 1.0))
            lbl_item.setPos((t_dbm2 + t_dbm3) / 2,
                            float(np.max(hr_values[hr_mask])) if np.any(hr_mask) else 80)
            plot1.addItem(lbl_item)

        # ===== 5-cycle average window =====
        t_avg_s = find_marker('start 5-cycle')
        t_avg_e = find_marker('end 5-cycle')
        if t_avg_s and t_avg_e:
            region = pg.LinearRegionItem(
                values=(t_avg_s, t_avg_e), orientation='vertical',
                brush=pg.mkBrush(255, 243, 205, 100), pen=pg.mkPen(None), movable=False)
            region.setZValue(-9)
            plot1.addItem(region)

        # ===== RSA: detect local max/min pairs within DBM2→DBM3 =====
        rsa_cycles = []
        if t_dbm2 and t_dbm3 and len(hr_times):
            C_INSP = '#B71C1C'   # dark red — HR max (inspiration)
            C_EXP  = '#1A237E'   # dark blue — HR min (expiration)

            # Work only within guided breathing window
            m_win = (hr_times >= t_dbm2) & (hr_times <= t_dbm3)
            t_win = hr_times[m_win]
            v_win = hr_values[m_win]

            if len(t_win) > 4:
                # Find all local maxima — must be > neighbours within ~3s window
                # Use simple 3-point peak detection with minimum prominence
                from scipy.signal import find_peaks
                min_dist_samples = max(1, int(4.0 / float(np.median(np.diff(t_win)))))
                peaks_i,  _ = find_peaks( v_win, distance=min_dist_samples, prominence=3.0)
                troughs_i, _ = find_peaks(-v_win, distance=min_dist_samples, prominence=3.0)

                print(f"DEBUG peaks at: {[round(float(t_win[i]),1) for i in peaks_i]}")
                print(f"DEBUG troughs at: {[round(float(t_win[i]),1) for i in troughs_i]}")

                # Pair each peak with the nearest trough that comes AFTER it
                used_troughs = set()
                cycle_n = 1
                for pi in peaks_i:
                    t_peak = float(t_win[pi])
                    v_peak = float(v_win[pi])
                    # Find first unused trough after this peak
                    best_ti = None
                    for ti in troughs_i:
                        if ti > pi and ti not in used_troughs:
                            best_ti = ti
                            break
                    if best_ti is None:
                        continue
                    used_troughs.add(best_ti)
                    t_trough = float(t_win[best_ti])
                    v_trough = float(v_win[best_ti])
                    rsa_val  = v_peak - v_trough

                    rsa_cycles.append({
                        'cycle':  cycle_n,
                        'max_t':  t_peak,
                        'max_v':  v_peak,
                        'min_t':  t_trough,
                        'min_v':  v_trough,
                        'rsa':    rsa_val,
                    })

                    # Dots on HR subplot
                    plot1.addItem(pg.ScatterPlotItem(
                        x=[t_peak], y=[v_peak], size=10, symbol='t',
                        pen=pg.mkPen(C_INSP, width=1.5),
                        brush=pg.mkBrush(C_INSP)))
                    plot1.addItem(pg.ScatterPlotItem(
                        x=[t_trough], y=[v_trough], size=10, symbol='t1',
                        pen=pg.mkPen(C_EXP, width=1.5),
                        brush=pg.mkBrush(C_EXP)))

                    cycle_n += 1

            print(f"DEBUG RSA cycles found: {len(rsa_cycles)}")
            for c in rsa_cycles:
                print(f"  Cycle {c['cycle']}: max={c['max_v']:.1f}@{c['max_t']:.1f}  "
                      f"min={c['min_v']:.1f}@{c['min_t']:.1f}  RSA={c['rsa']:.1f}")

            # ===== Populate QTableWidget =====
            if rsa_cycles:
                from pyqtgraph.Qt import QtWidgets
                valid = [c for c in rsa_cycles if c['rsa']]

                # Select top-6 by ΔHR (Novak/Rasmussen protocol)
                N_SELECT = 6
                top6 = set(
                    c['cycle'] for c in
                    sorted(valid, key=lambda x: x['rsa'], reverse=True)[:N_SELECT]
                )
                top6_cycles = [c for c in valid if c['cycle'] in top6]
                avg_rsa_all  = float(np.mean([c['rsa'] for c in valid])) if valid else 0.0
                avg_rsa_top6 = float(np.mean([c['rsa'] for c in top6_cycles])) if top6_cycles else 0.0

                # 7 columns: # | HR max | t max | HR min | t min | ΔHR | ✓
                table_widget.setColumnCount(7)
                table_widget.setHorizontalHeaderLabels(
                    ['#', 'HR max', 't max (s)', 'HR min', 't min (s)', 'ΔHR', 'Top 6'])
                table_widget.horizontalHeader().setSectionResizeMode(
                    6, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)

                def _cell(txt, color=None, bold=False, bg=None):
                    item = QtWidgets.QTableWidgetItem(str(txt))
                    item.setTextAlignment(pg.QtCore.Qt.AlignmentFlag.AlignCenter)
                    if color:
                        item.setForeground(pg.mkColor(color))
                    if bold:
                        f = item.font(); f.setBold(True); item.setFont(f)
                    if bg:
                        item.setBackground(pg.mkColor(bg))
                    return item

                # +2 rows: all-mean + top6-mean
                table_widget.setRowCount(len(rsa_cycles) + 2)

                for ri, c in enumerate(rsa_cycles):
                    selected = c['cycle'] in top6
                    bg = '#F1F8E9' if selected else None
                    table_widget.setItem(ri, 0, _cell(c['cycle'],             bg=bg))
                    table_widget.setItem(ri, 1, _cell(f"{c['max_v']:.1f}", '#8B0000', bg=bg))
                    table_widget.setItem(ri, 2, _cell(f"{c['max_t']:.1f}",             bg=bg))
                    table_widget.setItem(ri, 3, _cell(f"{c['min_v']:.1f}", '#1A237E', bg=bg))
                    table_widget.setItem(ri, 4, _cell(f"{c['min_t']:.1f}",             bg=bg))
                    table_widget.setItem(ri, 5, _cell(f"{c['rsa']:.1f}",               bg=bg))
                    chk = '✓' if selected else ''
                    table_widget.setItem(ri, 6, _cell(chk, '#2E7D32' if selected else None,
                                                      bold=selected, bg=bg))

                # Mean row — all cycles (grey)
                mr_all = len(rsa_cycles)
                mean_max_all = np.mean([c['max_v'] for c in valid])
                mean_min_all = np.mean([c['min_v'] for c in valid])
                BG_ALL  = '#F5F5F5'
                table_widget.setItem(mr_all, 0, _cell(f'Mean (n={len(valid)})', bold=True, bg=BG_ALL))
                table_widget.setItem(mr_all, 1, _cell(f"{mean_max_all:.1f}", '#8B0000', bold=True, bg=BG_ALL))
                table_widget.setItem(mr_all, 2, _cell('', bg=BG_ALL))
                table_widget.setItem(mr_all, 3, _cell(f"{mean_min_all:.1f}", '#1A237E', bold=True, bg=BG_ALL))
                table_widget.setItem(mr_all, 4, _cell('', bg=BG_ALL))
                table_widget.setItem(mr_all, 5, _cell(f"{avg_rsa_all:.1f}", '#555555', bold=True, bg=BG_ALL))
                table_widget.setItem(mr_all, 6, _cell('', bg=BG_ALL))

                # Mean row — top-6 (green, clinical result)
                mr_top = len(rsa_cycles) + 1
                mean_max_top = np.mean([c['max_v'] for c in top6_cycles]) if top6_cycles else 0.0
                mean_min_top = np.mean([c['min_v'] for c in top6_cycles]) if top6_cycles else 0.0
                BG_TOP  = '#E8F5E9'
                n_sel = min(N_SELECT, len(valid))
                table_widget.setItem(mr_top, 0, _cell(f'Mean (top {n_sel})', bold=True, bg=BG_TOP))
                table_widget.setItem(mr_top, 1, _cell(f"{mean_max_top:.1f}", '#8B0000', bold=True, bg=BG_TOP))
                table_widget.setItem(mr_top, 2, _cell('', bg=BG_TOP))
                table_widget.setItem(mr_top, 3, _cell(f"{mean_min_top:.1f}", '#1A237E', bold=True, bg=BG_TOP))
                table_widget.setItem(mr_top, 4, _cell('', bg=BG_TOP))
                table_widget.setItem(mr_top, 5, _cell(f"{avg_rsa_top6:.1f}", '#1B5E20', bold=True, bg=BG_TOP))
                table_widget.setItem(mr_top, 6, _cell('✓', '#2E7D32', bold=True, bg=BG_TOP))

                print(f"DEBUG RSA mean ΔHR all={avg_rsa_all:.1f}  top{n_sel}={avg_rsa_top6:.1f} bpm")

                # Store for export
                _export_rsa = dict(
                    cycles=rsa_cycles, valid=valid, top6=top6,
                    top6_cycles=top6_cycles, n_sel=n_sel,
                    avg_rsa_all=avg_rsa_all, avg_rsa_top6=avg_rsa_top6,
                    mean_max_all=mean_max_all, mean_min_all=mean_min_all,
                    mean_max_top=mean_max_top, mean_min_top=mean_min_top,
                )
            else:
                _export_rsa = None

        plot1.disableAutoRange()
        if t_dbm2 and t_dbm3:
            plot1.setXRange(t_dbm2 - 10, t_dbm3 + 10, padding=0)

        # Y range with extra headroom for label
        if np.any(hr_mask):
            y_min = float(np.min(hr_values[hr_mask]))
            y_max = float(np.max(hr_values[hr_mask]))
            plot1.setYRange(y_min - 5, y_max + 8, padding=0)

        plot1.getViewBox().setAutoPan(x=False, y=False)
        plot1.getViewBox().setAutoVisible(x=False, y=False)

        # ===== EXPORT PNG + XLSX =====
        if output_dir and _export_rsa is not None:
            import os, pathlib
            from pyqtgraph.Qt import QtWidgets, QtCore, QtGui

            out = pathlib.Path(output_dir)
            folder_prefix = out.name  # e.g. "2026-02-02_10.33.58"

            # --- PNG (HR plot only) ---
            png_path = out / f"{folder_prefix}_deep_breathing_results.png"
            from pyqtgraph.exporters import ImageExporter
            import pyqtgraph.Qt as pgQt
            pgQt.QtWidgets.QApplication.processEvents()
            exporter = ImageExporter(plot1)
            exporter.export(str(png_path))
            print(f"DEBUG: saved PNG → {png_path}")

            # --- XLSX ---
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.utils import get_column_letter
            except ImportError:
                import subprocess, sys
                subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                                       'openpyxl', '--break-system-packages', '-q'])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deep Breathing RSA"

            # Styles
            hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill  = PatternFill('solid', fgColor='1F4E79')
            ctr       = Alignment(horizontal='center', vertical='center')
            bold_blk  = Font(name='Arial', bold=True, size=11)
            norm      = Font(name='Arial', size=11)
            fill_green= PatternFill('solid', fgColor='E8F5E9')
            fill_grey = PatternFill('solid', fgColor='F5F5F5')
            fill_sel  = PatternFill('solid', fgColor='F1F8E9')
            fill_alt  = PatternFill('solid', fgColor='EEF2F7')
            red_font  = Font(name='Arial', size=11, color='8B0000')
            blue_font = Font(name='Arial', size=11, color='1A237E')
            red_bold  = Font(name='Arial', bold=True, size=11, color='8B0000')
            blue_bold = Font(name='Arial', bold=True, size=11, color='1A237E')
            grey_bold = Font(name='Arial', bold=True, size=11, color='555555')
            grn_bold  = Font(name='Arial', bold=True, size=11, color='1B5E20')
            chk_font  = Font(name='Arial', bold=True, size=11, color='2E7D32')
            thin      = Side(style='thin', color='CCCCCC')
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _wc(ws, row, col, val, font=None, fill=None, align=None, border=None):
                c = ws.cell(row=row, column=col, value=val)
                if font:   c.font   = font
                if fill:   c.fill   = fill
                if align:  c.alignment = align
                if border: c.border = border
                return c

            # Title
            ws.merge_cells('A1:G1')
            _wc(ws, 1, 1, 'Deep Breathing — RSA Analysis',
                font=Font(name='Arial', bold=True, size=13),
                align=Alignment(horizontal='center', vertical='center'))
            ws.row_dimensions[1].height = 22

            # Header row
            headers = ['#', 'HR max (bpm)', 't max (s)', 'HR min (bpm)', 't min (s)', 'ΔHR (bpm)', 'Top 6']
            for ci, h in enumerate(headers, 1):
                _wc(ws, 2, ci, h, font=hdr_font, fill=hdr_fill,
                    align=ctr, border=border)
            ws.row_dimensions[2].height = 18

            # Data rows
            d = _export_rsa
            for ri, c in enumerate(d['cycles']):
                row = ri + 3
                selected = c['cycle'] in d['top6']
                bg = fill_sel if selected else (fill_alt if ri % 2 else None)
                _wc(ws, row, 1, c['cycle'],    font=bold_blk if selected else norm, fill=bg, align=ctr, border=border)
                _wc(ws, row, 2, round(c['max_v'],1), font=red_font,  fill=bg, align=ctr, border=border)
                _wc(ws, row, 3, round(c['max_t'],1), font=norm,       fill=bg, align=ctr, border=border)
                _wc(ws, row, 4, round(c['min_v'],1), font=blue_font,  fill=bg, align=ctr, border=border)
                _wc(ws, row, 5, round(c['min_t'],1), font=norm,       fill=bg, align=ctr, border=border)
                _wc(ws, row, 6, round(c['rsa'],1),   font=norm,       fill=bg, align=ctr, border=border)
                _wc(ws, row, 7, '✓' if selected else '', font=chk_font if selected else norm,
                    fill=bg, align=ctr, border=border)

            n_data = len(d['cycles'])

            # Mean (all) row
            mr_all = n_data + 3
            _wc(ws, mr_all, 1, f"Mean (n={len(d['valid'])})", font=bold_blk, fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 2, round(d['mean_max_all'],1),    font=red_bold,  fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 3, '',                             font=norm,      fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 4, round(d['mean_min_all'],1),    font=blue_bold, fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 5, '',                             font=norm,      fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 6, round(d['avg_rsa_all'],1),     font=grey_bold, fill=fill_grey, align=ctr, border=border)
            _wc(ws, mr_all, 7, '',                             font=norm,      fill=fill_grey, align=ctr, border=border)

            # Mean (top N) row — clinical result
            mr_top = n_data + 4
            _wc(ws, mr_top, 1, f"Mean (top {d['n_sel']})",   font=bold_blk, fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 2, round(d['mean_max_top'],1),   font=red_bold,  fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 3, '',                            font=norm,      fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 4, round(d['mean_min_top'],1),   font=blue_bold, fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 5, '',                            font=norm,      fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 6, round(d['avg_rsa_top6'],1),   font=grn_bold,  fill=fill_green, align=ctr, border=border)
            _wc(ws, mr_top, 7, '✓',                           font=chk_font,  fill=fill_green, align=ctr, border=border)

            # Column widths — fit to content
            col_min = [6, 10, 10, 12, 10, 10, 7]
            for ci in range(1, 8):
                max_len = col_min[ci - 1]
                for row in ws.iter_rows(min_col=ci, max_col=ci):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)) + 2)
                ws.column_dimensions[get_column_letter(ci)].width = max_len

            # Embed PNG at I1, 40% larger than native render size
            if png_path.exists():
                from PIL import Image as PILImage
                with PILImage.open(str(png_path)) as im:
                    nw, nh = im.size
                xl_img = XLImage(str(png_path))
                xl_img.width  = int(nw * 0.8)
                xl_img.height = int(nh * 0.8)
                ws.add_image(xl_img, 'I1')

            xlsx_path = out / f"{folder_prefix}_deep_breathing_results.xlsx"
            wb.save(str(xlsx_path))
            print(f"DEBUG: saved XLSX → {xlsx_path}")

        print("DEBUG: create_deep_breathing_plot complete")